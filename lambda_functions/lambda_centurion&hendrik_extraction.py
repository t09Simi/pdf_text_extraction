import json
from datetime import datetime
import boto3
import re
import pdfplumber
from io import BytesIO
from openpyxl import load_workbook


s3 = boto3.client('s3')
lambda_client = boto3.client('lambda')
retries = 3


def move_pdf_to_out_bucket(source_bucket, object_key, file_content, file_folder):
    target_bucket = 'pdf-out-bucket'
    # Get the current date in the format YYYY-MM-DD
    current_date = datetime.now().strftime('%Y-%m-%d')

    # Check if the folder with the current date exists in the target S3 bucket
    target_folder_key = f'{file_folder}/{current_date}/{object_key}'

    # Check if the folder exists
    existing_objects = s3.list_objects(Bucket=target_bucket, Prefix=f'{file_folder}/{current_date}/')

    if 'Contents' not in existing_objects:
        # Folder doesn't exist, create it
        s3.put_object(Body='', Bucket=target_bucket, Key=f'{file_folder}/{current_date}/')

    # Upload the file to the target S3 bucket and folder
    s3.put_object(Body=file_content, Bucket=target_bucket, Key=target_folder_key)
    print(f"File moved to the target bucket: {target_bucket}, Folder: {file_folder}/{current_date}")

    # Delete the original file from the source bucket (uncomment if needed)
    s3.delete_object(Bucket=source_bucket, Key=object_key)
    print(f"File deleted from source bucket: {source_bucket}, file: {object_key}")


def invoke_excel_management_lambda(source_bucket, object_key, file_content, extracted_data, client, filename, page_errors):
    # Payload to pass to the second Lambda function
    global retries
    payload = {
        'extracted_data': extracted_data,
        'client': client,
        'filename': filename,
        'page_errors': page_errors
    }

    # Invoke the second Lambda function asynchronously
    response = lambda_client.invoke(
        FunctionName='excel_management',
        InvocationType='Event',
        Payload=json.dumps(payload)
    )

    # Optionally, you can check the response for success or handle errors
    status_code = response['StatusCode']
    if status_code == 202:
        print(f" Lambda function: excel_management invoked successfully.")
        move_pdf_to_out_bucket(source_bucket, object_key, file_content, "Success")
    else:
        print(f"Error invoking Lambda function: excel_management. Status code: {status_code}")
        if retries:
            retries -= 1
            print("Retrying to invoke the Lambda function: excel_management")
            invoke_excel_management_lambda(source_bucket, object_key, file_content, extracted_data, client, filename, page_errors)
        else:
            move_pdf_to_out_bucket(source_bucket, object_key, file_content, "Failure")


def get_manufacture_model(description: str):
    # Specify the S3 bucket and file name
    excel_bucket = 'resources-and-extraction-data'
    excel_file_key = 'Full_list_of_Manufacturers_and_Models.xlsx'

    # Download the Excel file from S3
    excel_file_obj = s3.get_object(Bucket=excel_bucket, Key=excel_file_key)
    excel_file_content = excel_file_obj['Body'].read()

    workbook = load_workbook(filename=BytesIO(excel_file_content), read_only=True)
    manufacturer_sheet = workbook['Manufacture']
    model_sheet =workbook['Model']
    description_keyword =description.lower().split()
    manufacture, model = "",""
    for row in manufacturer_sheet.iter_rows(min_row=2, values_only=True):
        keyword = row[0]
        if keyword and str(keyword).lower() in description_keyword:
            manufacture = keyword
            break
    for row in model_sheet.iter_rows(min_row=2, values_only=True):
        keyword = row[0]
        if keyword and str(keyword).lower() in description_keyword:
            model = keyword
            break
    return manufacture, model


def extract_quantity(text):
    match = re.match(r'^\s*(\d+)\s+\d*', text)
    if match:
        return int(match.group(1))
    else:
        return None


def process_swl(swl: str):
    pattern = r'^(\d+(?:\.\d+)?)([a-zA-Z]+)?\s*(.*)$'

    # Match the pattern
    match = re.match(pattern, swl)

    if match:
        value_part = match.group(1)
        unit_part = match.group(2)
        note_part = match.group(3)
    else:
        value_part = None
        unit_part = None
        note_part = swl
    units_map = {"kgs": "kg"}
    if unit_part in units_map:
        unit_part = units_map[unit_part]
    # Check if part 2 is a unit type or not
    units = ["kg", "g", "lb", "ton", "t", "m", "cm", "mm", "ft", "in", "m²", "cm²", "mm²", "ft²", "in²", "m³", "cm³",
             "mm³",
             "ft³", "in³", "km/h", "mph", "m/s", "kph", "°C", "°F", "°K", "bar", "atm", "Pa", "kPa", "psi", "N", "J",
             "W", "A", "V", "F", "Ω", "S", "H", "Hz", "C", "Bq", "Gy", "Sv", "cd", "lm", "lx", "B", "mol", "unit", "te"]
    if unit_part and unit_part not in units:
        note_part = unit_part + " " + note_part if note_part else unit_part
        unit_part = None

    return value_part, unit_part, note_part



def get_identification_parts_list(input_string: str, quantity: int):
    numeric_part = ''.join(filter(str.isdigit, input_string))
    part_list = list()
    if input_string[0].isdigit():
        alpha_part = input_string[len(numeric_part):]
        for i in range(0, quantity):
            part_list.append(f"{int(numeric_part) + i}{alpha_part}")
    else:
        alpha_part = input_string[:len(input_string)-len(numeric_part)]
        for i in range(0, quantity):
            part_list.append(f"{alpha_part}{int(numeric_part) + i}")
    return part_list


def get_identification_number_list(identification_numbers: str, quantity: int):
    identification_number_list = []
    if "to" in identification_numbers.lower():
        # identification_number_first_part = D971-1 or MGL1
        identification_number_first_part = identification_numbers.split("to")[0].strip()
        # print(identification_number_first_part)
        # example: D971-1
        if "-" in identification_number_first_part:
            # first_part = D971, second_part = 1
            first_part, second_part = identification_number_first_part.split('-')
            # print(first_part, second_part)
            second_part_list = get_identification_parts_list(second_part, quantity)
            # print(second_part_list)
            identification_number_list = list()
            for second_part in second_part_list:
                identification_number_list.append(f"{first_part}-{second_part}")
        else:
            # example: MGL1
            identification_number_list = get_identification_parts_list(identification_number_first_part, quantity)
    elif re.search(r'x(\d+)', identification_numbers):
        # print(identification_numbers)
        identification_number_list = list()
        for num in identification_numbers.split(','):
            id_number = num.split('x')[0].strip()
            for i in range(len(id_number)-1, -1, -1):
                if id_number[i].isdigit():
                    id_number= id_number[:i+1]
                    break
            count = int(''.join(filter(str.isdigit, num.split('x')[1].strip())))
            for i in range(1,count+1):
                identification_number_list.append(f"{id_number}-{i}")
    elif "," in identification_numbers:
        identification_number_list = identification_numbers.split(',')

    # print(identification_number_list)
    return identification_number_list


def extraction_centurion_pdf(source_bucket, object_key):
    try:
        print("<-------------extracting centurion pdf------------>")
        pdf_file = s3.get_object(Bucket=source_bucket, Key=object_key)
        file_content = pdf_file['Body'].read()
        pdf_doc = pdfplumber.open(BytesIO(file_content))
        extraction_info = dict()
        page_errors = dict()
        for i, page in enumerate(pdf_doc.pages):
            try:
                text = page.extract_text()
                page = pdf_doc.pages[i]
                if "Centurion" in text and "Hendrik" not in text:
                    if page.extract_tables():
                        print("page number:", i)
                        page_tables = page.extract_tables()
                        first_table = page_tables[0]
                        # data1:Report Number / Date of Examination / Ref No
                        table_data1 = first_table[0][13]
                        # data2: Identify Company keywords
                        table_data2 = first_table[1]
                        # data3: Identify Text keywords
                        table_data3 = first_table[2]
                        # data4: Provide related Value
                        table_data4 = first_table[4]
                        # data5 ID number Value
                        table_data5 = first_table[5]
                        errors = list()
                        identification_number_list = list()
                        if "Quantity & Description of Equipment, Serial Numbers" in table_data3[0]:

                            id_numbers, description, mnfer, wwl, next_thorough = None, None, None, None, None
                            quantity = 1

                            for index in range(0, len(table_data3)):
                                try:
                                    if table_data3[index] is None:
                                        continue
                                    text_to_compare = table_data3[index].lower()
                                    if not description and "description" in text_to_compare:
                                        description = table_data4[0].replace('\n', ' ')
                                        serial_numbers = table_data5[index].split(":")[-1].strip()
                                        mnfer = table_data4[4].strip()

                                    elif not wwl and "working" in text_to_compare:
                                        wwl = table_data4[index].strip()
                                    elif not next_thorough and "next" in text_to_compare:
                                        date_string = table_data4[index].strip()
                                        date_obj = datetime.strptime(date_string, "%d/%m/%Y")
                                        next_thorough = date_obj.strftime("%d/%m/%Y")
                                    elif not id_numbers and "certificate" in text_to_compare:
                                        id_numbers = table_data4[index].strip()
                                except Exception as e:
                                    print("Error extracting value from page:", e)

                            if id_numbers:
                                page_info = dict()
                                if description:
                                    try:
                                        item_description = description
                                        if not item_description:
                                            errors.append("Item Description not found")
                                        else:
                                            page_info["Item Description"] = description.split(':')[0]
                                    except Exception as e:
                                        errors.append(e)
                                    try:
                                        manufacturer, model = get_manufacture_model(description)
                                        manufacturer = mnfer
                                        if not manufacturer:
                                            errors.append("Manufacturer not found")
                                        else:
                                            page_info["Manufacturer"] = manufacturer
                                        if not model:
                                            errors.append("Model not found")
                                        else:
                                            page_info["Model"] = model
                                    except Exception as e:
                                        errors.append(e)

                                if wwl:
                                    try:
                                        swl_value, swl_unit, swl_note = process_swl(wwl)
                                        if not swl_value:
                                            errors.append("SWL Value not found")
                                        else:
                                            page_info["SWL Value"] = swl_value
                                        if not swl_unit:
                                            errors.append("SWL Unit not Found")
                                        else:
                                            page_info["SWL Unit"] = swl_unit
                                        page_info["SWL Note"] = swl_note
                                    except Exception as e:
                                        errors.append(e)
                                else:
                                    errors.append("SWL not found in this page.")
                                page_info["Next Inspection Due Date"] = next_thorough
                                # report_number, date_of_examination, job_number, next_date_of__examination = None, None, None, None
                                table_data1_mapping = dict()
                                table_data1 = table_data1.splitlines()

                                for data in table_data1:
                                    data_list = data.split(':', 1)
                                    if len(data_list) == 2:
                                        key, value = data_list
                                        formattted_key = key.lower().replace(" ", "").replace("/", "").replace(".", "")
                                        table_data1_mapping[formattted_key] = value.strip()

                                page_info["Provider Identification"] = table_data1_mapping["custrefpono"]
                                page_info["Certificate No"] = page_info["Provider Identification"]
                                page_info["Previous Inspection"] = table_data1_mapping["dateofexamination"]

                                id_numbers = serial_numbers

                                if quantity == 1:
                                    identification_number_list.append(id_numbers)

                                for identification_number in identification_number_list:
                                    extraction_info[identification_number] = page_info

                            else:
                                print("No identification error")

                        elif "Qty, Description of Equipment, Serial Numbers" in table_data3[0]:
                            quantity, id_numbers, description, mnfer, wwl, next_thorough = None, None, None, None, None, None
                            for index in range(0, len(table_data3)):
                                if table_data3[index] is None:
                                    continue
                                text_to_compare = table_data3[index].lower()
                                if not description and "description" in text_to_compare:
                                    description = table_data4[index].replace('\n', ' ')
                                    serial_numbers = table_data5[index].split(":")[-1].strip()
                                    quantity = extract_quantity(table_data4[index])
                                    mnfer = table_data4[4].strip()
                                elif not wwl and "working" in text_to_compare:
                                    wwl = table_data4[index].strip()
                                elif not next_thorough and "next" in text_to_compare:
                                    date_string = table_data4[index].strip()
                                    date_obj = datetime.strptime(date_string, "%d/%m/%Y")
                                    next_thorough = date_obj.strftime("%d/%m/%Y")
                                elif not id_numbers and "certificate" in text_to_compare:
                                    id_numbers = table_data4[index].strip()

                            if id_numbers:
                                page_info = dict()
                                if description:
                                    page_info["Item Description"] = description.split(':')[0]
                                    manufacturer, model = get_manufacture_model(description)
                                    manufacturer = mnfer
                                    page_info["Manufacturer"] = manufacturer
                                    page_info["Model"] = model
                                if wwl:
                                    try:
                                        swl_value, swl_unit, swl_note = process_swl(wwl)
                                        if not swl_value:
                                            errors.append("SWL Value not found")
                                        else:
                                            page_info["SWL Value"] = swl_value
                                        if not swl_unit:
                                            errors.append("SWL Unit not Found")
                                        else:
                                            page_info["SWL Unit"] = swl_unit
                                        page_info["SWL Note"] = swl_note
                                    except Exception as e:
                                        errors.append(e)
                                else:
                                    errors.append("SWL not found in this page.")
                                page_info["Next Inspection Due Date"] = next_thorough
                                # report_number, date_of_examination, job_number, next_date_of__examination = None, None, None, None
                                table_data1_mapping = dict()
                                table_data1 = table_data1.splitlines()

                                for data in table_data1:
                                    data_list = data.split(':', 1)
                                    if len(data_list) == 2:
                                        key, value = data_list
                                        formattted_key = key.lower().replace(" ", "").replace("/", "").replace(".", "")
                                        table_data1_mapping[formattted_key] = value.strip()

                                page_info["Provider Identification"] = table_data1_mapping["custrefpono"]
                                page_info["Certificate No"] = page_info["Provider Identification"]
                                page_info["Previous Inspection"] = table_data1_mapping["dateofexamination"]

                                id_numbers = serial_numbers

                                if quantity > 1:
                                    identification_number_list = get_identification_number_list(id_numbers, quantity)

                                for identification_number in identification_number_list:
                                    extraction_info[identification_number] = page_info

                                # print(identification_numbers, page_info)
                            else:
                                print("No identification error")


                elif "Hendrik" in text:
                    page = pdf_doc.pages[i]
                    text = page.extract_text()

                    # data1: Certificate No.
                    certificate_no = None
                    certificate_match = re.search(r'Certificate No\. :\s*(\d+)', text)
                    if certificate_match:
                        print("page number:", i)
                        certificate_no = certificate_match.group(1)

                    page_tables = page.extract_tables()
                    first_table = page_tables[0]
                    # data2: wwl
                    table_data1 = first_table[3]
                    # data3: description, manufacturer
                    table_data2 = first_table[9]
                    table_data4 = first_table[10]
                    # data4: previous inspection
                    table_data3 = first_table[14]
                    # data5: ID Number
                    table_data5 = first_table[2]

                    identification_number_list = list()

                    id_numbers, wwl, pre_date = None, None, None
                    id_data = table_data5[0]
                    id_match = re.search(r'\)\s*([^\s]+)', id_data)
                    if id_match:
                        id_numbers = id_match.group(1)

                    wwl_data = table_data1[0]
                    wwl_match = re.search(r'(\d+(\.\d+)?\s*t)', wwl_data)
                    if wwl_match:
                        wwl = wwl_match.group(1)

                    pre_data = table_data3[0]
                    pre_match = re.search(r'(\d{2}-\d{2}-\d{4})', pre_data)
                    if pre_match:
                        pre_value = pre_match.group(1)
                        pre_date = pre_value.replace('-', '/')

                    description, next_thorough, provider, manufacturer, model = None, None, None, None, None
                    quantity = 1

                    for index in range(0, len(table_data2)):
                        if table_data2[index] is None:
                            continue
                        text_to_compare = table_data2[index].lower()
                        if not description and 'description' in text_to_compare:
                            description = table_data4[index].replace('\n', ' ')
                            manufacturer_data = table_data4[2]

                    if id_numbers:
                        page_info = dict()
                        if description:
                            page_info["Item Description"] = description.split(':')[0]
                            try:
                                manufacturer, model = get_manufacture_model(manufacturer_data)
                                model = get_manufacture_model(description)
                                if not manufacturer:
                                    errors.append("Manufacturer not found")
                                else:
                                    page_info["Manufacturer"] = manufacturer
                                if not model:
                                    errors.append("Model not found")
                                else:
                                    page_info["Model"] = model
                            except Exception as e:
                                errors.append(e)
                            if wwl:
                                try:
                                    swl_value, swl_unit, swl_note = process_swl(wwl)
                                    if not swl_value:
                                        errors.append("SWL Value not found")
                                    else:
                                        page_info["SWL Value"] = swl_value
                                    if not swl_unit:
                                        errors.append("SWL Unit not Found")
                                    else:
                                        page_info["SWL Unit"] = swl_unit
                                    page_info["SWL Note"] = swl_note
                                except Exception as e:
                                    errors.append(e)
                            else:
                                errors.append("SWL not found in this page.")
                            page_info["Certificate No"] = certificate_no
                            page_info["Next Inspection Due Date"] = next_thorough
                            page_info["Provider Identification"] = provider
                            page_info["Previous Inspection"] = pre_date

                            if quantity == 1:
                                identification_number_list.append(id_numbers)

                            for identification_number in identification_number_list:
                                extraction_info[identification_number] = page_info
                    else:
                        print("No identification error")
                else:
                    print("No verified company found")
            except Exception as e:
                page_errors[i+1] = "Error" + str(e) + "occurred while processing the page:"
                print("Error", e, "occurred while processing the page: ", i)
        invoke_excel_management_lambda(source_bucket, object_key, file_content, extraction_info, "Centurion", object_key.replace("pdf", "xlsx"), page_errors)
    except Exception as e:
        print("An error occurred while processing in the pdf:", e)
        move_pdf_to_out_bucket(source_bucket, object_key, file_content, "Failure")


def lambda_handler(event, context):
    # Parse the payload from the event
    try:
        source_bucket = event['source_bucket']
        object_key = event['object_key']
        print(f"Received payload from the first Lambda function. Source bucket: {source_bucket}, Object key: {object_key}")
        extraction_centurion_pdf(source_bucket, object_key)
    except Exception as e:
        print("An error occurred while decoding source bucket and object key:", e)

