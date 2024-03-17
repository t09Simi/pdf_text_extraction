import json
from datetime import datetime, timedelta
import boto3
import re
import pdfplumber
from io import BytesIO
from openpyxl import load_workbook

s3 = boto3.client('s3')
lambda_client = boto3.client('lambda')


def move_pdf_to_out_bucket(source_bucket, object_key, file_content):
    target_bucket = 'pdf-out-bucket'
    # Get the current date in the format YYYY-MM-DD
    current_date = datetime.now().strftime('%Y-%m-%d')

    # Check if the folder with the current date exists in the target S3 bucket
    target_folder_key = f'{current_date}/{object_key}'

    # Check if the folder exists
    existing_objects = s3.list_objects(Bucket=target_bucket, Prefix=f'{current_date}/')

    if 'Contents' not in existing_objects:
        # Folder doesn't exist, create it
        s3.put_object(Body='', Bucket=target_bucket, Key=f'{current_date}/')

    # Upload the file to the target S3 bucket and folder
    s3.put_object(Body=file_content, Bucket=target_bucket, Key=target_folder_key)
    print(f"File moved to the target bucket: {target_bucket}, Folder: {current_date}")

    # Delete the original file from the source bucket (uncomment if needed)
    s3.delete_object(Bucket=source_bucket, Key=object_key)
    print(f"File deleted from source bucket: {source_bucket}, file: {object_key}")


def invoke_excel_management_lambda(source_bucket, object_key, file_content, extracted_data, sheet_name):
    # Payload to pass to the second Lambda function
    payload = {
        'extracted_data': extracted_data,
        'sheet_name': sheet_name
    }

    # Invoke the second Lambda function asynchronously
    response = lambda_client.invoke(
        FunctionName='excel_management',
        InvocationType='Event',
        Payload=json.dumps(payload)
    )

    # Optionally, you can check the response for success or handle errors
    status_code = response['StatusCode']
    if status_code == 202:
        print(f" Lambda function: excel_management invoked successfully.")
        move_pdf_to_out_bucket(source_bucket, object_key, file_content)
    else:
        print(f"Error invoking Lambda function: excel_management. Status code: {status_code}")


def split_id_numbers_with_range(id_numbers):
    new_id_numbers = []
    error_statements = []
    for id_number in id_numbers:

        match = re.match(r'([A-Za-z]+)(\d+)-(\d+)', id_number)
        if match:
            prefix_alpha = match.group(1)
            range_start = int(match.group(2))
            range_end = int(match.group(3))
            if range_end < range_start:
                error_statement = f"Error: End of range ({range_end}) is less than start of range ({range_start}) for ID number {id_number}"
                error_statements.append(error_statement)
                continue
            for i in range(range_start, range_end + 1):
                new_id_numbers.append(f"{prefix_alpha}{i}")

        else:
            # Check if the second pattern matches
            match = re.match(r'([A-Za-z]+\d+/?\d*)/(\d+)-(\d+)', id_number)
            if match:
                prefix_alpha = match.group(1)
                prefix_numeric = match.group(2)
                range_start = int(prefix_numeric)
                range_end = int(match.group(3))
                if range_end < range_start:
                    error_statement = f"Error: End of range ({range_end}) is less than start of range ({range_start}) for ID number {id_number}"
                    error_statements.append(error_statement)
                    continue
                # Generate new id numbers based on the matched pattern
                for i in range(range_start, range_end + 1):
                    new_id_numbers.append(f"{prefix_alpha}/{i}")
            else:
                # If neither pattern matches, append the id_number as is
                new_id_numbers.append(id_number)
                continue
    return new_id_numbers


def get_manufacture_model(description: str):
    # Specify the S3 bucket and file name
    excel_bucket = 'extraction-and-resources-data'
    excel_file_key = 'Model Numbers with their Manufacturers.xlsx'

    # Download the Excel file from S3
    excel_file_obj = s3.get_object(Bucket=excel_bucket, Key=excel_file_key)
    excel_file_content = excel_file_obj['Body'].read()

    workbook = load_workbook(filename=BytesIO(excel_file_content), read_only=True)
    sheet = workbook['Sheet1']

    description_keywords = description.lower()
    manufacture, model = "", ""

    # Search for manufacturer in Manufacturer column
    for row in sheet.iter_rows(min_row=2, values_only=True):
        model_keyword = row[1].strip()

        if model_keyword and str(model_keyword).lower() in description_keywords:
            model = model_keyword
            manufacture = row[0]
            break

    return manufacture, model


def contains_keyword(first_row, keyword):
    if first_row is not None:
        row_text = ' '.join(cell if cell is not None else '' for cell in first_row)
        if row_text.lower().startswith(keyword.lower()):
            return True
    return False


def add_six_months(date_str):
    date_declaration_obj = datetime.strptime(date_str, '%d/%m/%Y')
    # Add 6 months to the date of declaration
    next_inspection_date_obj = date_declaration_obj + timedelta(days=6 * 30)
    # Convert the next inspection date back to a string in the same format
    next_inspection_date_str = next_inspection_date_obj.strftime('%d/%m/%Y')

    return next_inspection_date_str


def extract_first_integrated_pdf(source_bucket, object_key):
    print("<------------extracting first_integrated pdf------------>")
    pdf_file = s3.get_object(Bucket=source_bucket, Key=object_key)
    file_content = pdf_file['Body'].read()
    pdf_doc = pdfplumber.open(BytesIO(file_content))
    extraction_info = dict()
    for i in range(0, len(pdf_doc.pages)):
        # for page_number, page in enumerate(pdf_doc.pages()):
        page = pdf_doc.pages[i]
        page_tables = page.extract_tables()
        # print("page number:", i)
        # print("page tables:", page_tables)
        if not page_tables:
            print(f"No tables found on page {i}. Skipping...")
            continue

        first_row = page_tables[0][0]

        if contains_keyword(first_row, "Name & Address of employer for Whom the examination was made"):
            process_table_type1(page_tables, extraction_info)
        elif contains_keyword(first_row, "Date of Thorough Examination"):
            process_table_type2(page_tables[0], extraction_info)
        elif contains_keyword(first_row, "Name &AddressofManufacturer") or contains_keyword(first_row,
                                                                                            "Name & Address of Manufacturer"):
            process_table_type3(page_tables[0], extraction_info)
        else:
            print("No recognized table found on page", i)

    invoke_excel_management_lambda(source_bucket, object_key, file_content, extraction_info, "First Integrated")


def process_table_type1(page_tables, extraction_info):
    id_number = None
    id_pattern = re.compile(r'(\(\w+\)\s*)?[A-Z]{3}\d+')
    swl_pattern = re.compile(r'(\d+\.\d+)\s+(Tonnes|Kilos)\s', re.IGNORECASE)
    row_data = dict()
    report_ref_no, date_of_thorough_examination = None, None

    for page in page_tables:
        for index, row in enumerate(page):
            page_info = {}
            row_text = ' '.join(cell if cell is not None else '' for cell in row)

            if "Name & Address of employer" in row_text:
                continue  # Skip this row
                # Extract ID Number
            id_match = id_pattern.search(row_text)
            if id_match:
                id_number = id_match.group()
                # print("id_number",id_number)
                # page_info["Id Number"] = id_number

                # Extract SWL
                swl_match = swl_pattern.search(row_text)
                if swl_match:
                    swl_value = swl_match.group(1)
                    swl_units = swl_match.group(2)
                    swl_note = None
                    page_info["SWL Value"] = swl_value
                    page_info["SWL Unit"] = swl_units
                    page_info["SWL Note"] = swl_note

                # Extract Description between ID Number and SWL
                if id_match:
                    start_index = id_match.end()
                    if swl_match:
                        end_index = swl_match.start()
                    else:
                        end_index = -1
                    description = row_text[start_index:end_index].strip()
                    description = re.sub(r'\([^)]*\)\s*', '', description)
                    page_info["Item Description"] = description
                    # print("description", description)

                    # Get manufacture and model fro  description
                    manufacture, model = get_manufacture_model(description)
                    page_info["Manufacturer"] = manufacture
                    page_info["Model"] = model

                # Extract Date of Next Inspection
                next_inspection_pattern = re.compile(r'(\d{2}/\d{2}/\d{4})$', re.IGNORECASE)
                next_inspection_match = next_inspection_pattern.search(row_text)
                if next_inspection_match:
                    page_info["Next Inspection Due Date"] = next_inspection_match.group(1).strip()
                    # print("Next Inspection Date:", next_inspection_match)

                row_data[id_number] = page_info

    # Extract Date of Previous Inspection and Certificate Number
    for page in page_tables:
        for row in page:
            row_text = ' '.join(cell if cell is not None else '' for cell in row)

            # Extract Date of Previous Inspection
            keyword_date = "Date of thorough examination"
            if keyword_date.lower() in row_text.lower():
                date_pattern = re.compile(r'\b(\d{2}/\d{2}/\d{4})\b')
                previous_inspection_match = date_pattern.search(row_text)
                if previous_inspection_match:
                    date_of_thorough_examination = previous_inspection_match.group().strip()

            # Extract Report No
            keyword_report = "Report Ref No"
            if keyword_report.lower() in row_text.lower():
                report_pattern = re.compile(r'[A-Z]{3}/\d{6}/\d{5}')
                report_ref_match = report_pattern.search(row_text)
                if report_ref_match:
                    report_ref_no = report_ref_match.group().strip()

    # Assign Previous Inspection Date and Certificate No to all items in row_data
    for id_number, page_info in row_data.items():
        page_info["Previous Inspection"] = date_of_thorough_examination
        page_info["Certificate No"] = report_ref_no
        extraction_info[id_number] = page_info


def process_table_type2(table, extraction_info):
    page_info = {}
    id_number = None
    report_number_pattern = re.compile(r'Report\s*Number:', re.IGNORECASE)

    for row_outer in table:
        # print("row_outer", row_outer)
        for cell in row_outer:
            if cell:
                # Extract information based on conditions
                if "identification of the equipment" in cell:
                    parts = cell.split('\n')
                    if len(parts) >= 2:
                        description = parts[1].strip()
                        id_number = parts[2].strip()
                        # page_info["Id Number"] = id_number
                        page_info["Item Description"] = description
                        id_number_list = []
                        if "-" in id_number:
                            split_id_numbers = split_id_numbers_with_range([id_number])
                            id_number_list.extend(split_id_numbers)
                        else:
                            id_number_list.append(id_number)
                        for id_number in id_number_list:
                            extraction_info[id_number] = page_info

                        # Get manufacture and model from description
                        manufacture, model = get_manufacture_model(description)
                        page_info["Manufacturer"] = manufacture
                        page_info["Model"] = model

                elif "WLL" in cell:
                    parts = cell.split('\n')
                    if len(parts) >= 4:
                        wll = parts[3].strip()
                        swl_pattern = re.compile(r'(\d+(?:\.\d+)?)\s*(TONNE|TONNES)\b', re.IGNORECASE)
                        swl_match = swl_pattern.search(wll)
                        if swl_match:
                            swl_value = swl_match.group(1)
                            swl_units = swl_match.group(2)
                            swl_note = None
                            page_info["SWL Value"] = swl_value
                            page_info["SWL Unit"] = swl_units
                            page_info["SWL Note"] = swl_note

                elif report_number_pattern.search(cell):
                    certificate_number = report_number_pattern.sub('', cell).strip()
                    page_info["Certificate No"] = certificate_number
                elif "Date of Thorough" in cell:
                    previous_inspection = cell.replace("Date of Thorough Examination:", "").strip()
                    page_info["Previous Inspection"] = previous_inspection
                elif "Latest date by which next" in cell:
                    next_inspection = cell.replace(
                        "Latest date by which next thorough\nexamination must be carried out:", "").strip()
                    page_info["Next Inspection Due Date"] = next_inspection


def process_table_type3(table, extraction_info):
    page_info = dict()
    id_number = None
    for i, row in enumerate(table):
        if row:
            # Check for labels
            for j, cell in enumerate(row):
                if cell:
                    # Check for Id Number label
                    if "Id Number" in cell:
                        # Get the value from the next row
                        id_number_row = table[i + 1] if i + 1 < len(table) else None
                        if id_number_row:
                            id_number = id_number_row[j].strip() if j < len(id_number_row) else None
                            id_number_list = []
                            if "-" in id_number:
                                split_id_numbers = split_id_numbers_with_range([id_number])
                                id_number_list.extend(split_id_numbers)
                            else:
                                id_number_list.append(id_number)
                            for id_number in id_number_list:
                                extraction_info[id_number] = page_info
                    # Check for Description label
                    elif "Description" in cell:
                        parts = cell.split('\n')
                        if len(parts) >= 2:
                            description = parts[1].strip()
                            page_info["Item Description"] = description
                            # Get manufacture and model from description
                            manufacture, model = get_manufacture_model(description)
                            page_info["Manufacturer"] = manufacture
                            page_info["Model"] = model
                    # Check for WLL label
                    elif "WLL" in cell:
                        # Get the value from the next row
                        wll_row = table[i + 1] if i + 1 < len(table) else None
                        if wll_row:
                            wll = wll_row[j].strip() if j < len(wll_row) else None
                            swl_pattern = re.compile(r'(\d+(?:\.\d+)?)\s*(TONNE|TONNES)\b', re.IGNORECASE)
                            swl_match = swl_pattern.search(wll)
                            if swl_match:
                                swl_value = swl_match.group(1)
                                swl_units = swl_match.group(2)
                                page_info["SWL Value"] = swl_value
                                page_info["SWL Unit"] = swl_units
                    elif "Certificate Number" in cell:
                        certificate_row = table[i + 1] if i + 1 < len(table) else None
                        if certificate_row:
                            certificate_number = certificate_row[j].strip() if j < len(
                                certificate_row) else None
                            page_info["Certificate No"] = certificate_number
                    elif "Date of Declaration" in cell:
                        parts = cell.split('\n')
                        if len(parts) >= 3:
                            date_of_declaration = parts[3].strip()
                            # Use the function to add 6 months to the date
                            next_inspection_date = add_six_months(date_of_declaration)
                            page_info["Previous Inspection"] = date_of_declaration
                            page_info["Next Inspection Due Date"] = next_inspection_date


def lambda_handler(event, context):
    # Parse the payload from the event
    source_bucket = event['source_bucket']
    object_key = event['object_key']
    print(f"Received payload from the first Lambda function. Source bucket: {source_bucket}, Object key: {object_key}")
    extract_first_integrated_pdf(source_bucket, object_key)