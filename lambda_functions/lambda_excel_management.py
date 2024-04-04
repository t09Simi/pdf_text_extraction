import json
import boto3
from io import BytesIO
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side

s3 = boto3.client('s3')
retries = 3
topic_arn = ""


def send_sns(message, subject):
    try:
        client = boto3.client("sns")
        result = client.publish(TopicArn=topic_arn, Message=message, Subject=subject)
        if result['ResponseMetadata']['HTTPStatusCode'] == 200:
            print(result)
            print("Notification send successfully..!!!")
            return True
    except Exception as e:
        print("Error occured while publish notifications and error is : ", e)
        return True


def save_workbook_to_s3(workbook, bucket, key):
    # Save the modified workbook to bytes
    print("<--------------saving excel on the bucket------------------>")
    current_date = datetime.now().strftime('%Y-%m-%d')
    target_key = f'{current_date}/{key}'
     # Check if the folder exists
    existing_objects = s3.list_objects(Bucket=bucket, Prefix=f'{current_date}/')

    if 'Contents' not in existing_objects:
        # Folder doesn't exist, create it
        s3.put_object(Body='', Bucket=bucket, Key=f'{current_date}/')
    buffer = BytesIO()
    workbook.save(buffer)
    # Upload the modified Excel file back to S3, overwriting the original file
    s3.put_object(Body=buffer.getvalue(), Bucket=bucket, Key=target_key)


def create_excel(extracted_data: dict, filename: str, client: str, page_errors: dict, column_mapping: dict):
    print("<--------------Creating new excel------------------>")
    workbook = openpyxl.Workbook()  # Create a new Workbook

    # Create a sheet for extracted data
    sheet_data = workbook.active
    sheet_data.title = "Extraction Data"  # Set sheet name

    sheet_data['A1'] = "Rig-Ware import v2"
    sheet_data['B1'] = client
    sheet_data['C1'] = "CreateLocations=No"

    # Write column headers for extracted data sheet
    for header, column in column_mapping.items():
        cell = sheet_data[column + '2']
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align the text
        cell.border = Border(bottom=Side(border_style='thin'))  # Add a thin border at the bottom

        # Adjust column width to fit the header text
        column_width = max(len(header), 10)  # Set a minimum width of 10 characters
        sheet_data.column_dimensions[column].width = column_width

    # Write extracted data to the worksheet
    for row_idx, (key, data) in enumerate(extracted_data.items(), start=3):
        sheet_data.cell(row=row_idx, column=1, value=key)  # Write key in the first column
        for cell_name, value in data.items():
            column_name = column_mapping.get(cell_name)
            if column_name:
                sheet_data.cell(row=row_idx, column=ord(column_name) - 64, value=value)

    # Create a sheet for errors
    sheet_errors = workbook.create_sheet(title="Errors")  # Create a new worksheet
    sheet_errors.append(["Page No", "Error"])  # Write column headers

    # Write errors to the worksheet
    for key, value in page_errors.items():
        sheet_errors.append([key, value])  # Write key-value pairs as rows

    save_workbook_to_s3(workbook, 'excel-extraction-data', filename)
    # Close the workbook
    workbook.close()
    print("<-------------- Excel created successfully ------------------>")


def lambda_handler(event, context):
    global retries
    try:
        # Parse the payload from the event
        extracted_data = event['extracted_data']
        client = event['client']
        filename = event['filename']
        page_errors = event['page_errors']
        print(f"Received payload from other Lambda function. extracted_data: {extracted_data}, client: {client}, filename: {filename}")
        column_mapping = {
            "Id Number": "A",
            "RFID": "B",
            "Item Category": "C",
            "Item Description": "D",
            "Model": "E",
            "SWL Value": "F",
            "SWL Unit": "G",
            "SWL Note": "H",
            "Manufacturer": "I",
            "Certificate No": "J",
            "Location": "K",
            "Detailed Location ": "L",
            "Previous Inspection": "M",
            "Next Inspection Due Date": "N",
            "Fit For Purpose Y/N": "O",
            "Status": "P",
            "Provider Identification": "Q",
            "Errors": "R"
        }
        create_excel(extracted_data, filename, client, page_errors, column_mapping)
    except Exception as e:
        print(f"An error occurred in excel creation: {e}")
        if retries:
            retries -= 1
            lambda_handler(event, context)
        else:
            message = "Excel file creation is failed for the following file: " + filename.replace("xlsx", "pdf")
            subject = "An error occurred in excel creation"
            send_sns(message, subject)
