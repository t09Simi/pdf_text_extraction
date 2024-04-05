import re
import pdfplumber
from datetime import datetime, timedelta
from openpyxl import load_workbook
import excel_management


def split_id_numbers_with_range(id_numbers):
    new_id_numbers = []
    new_errors = list()
    for id_number in id_numbers:
        match = re.match(r'([A-Za-z]+)(\d+)-(\d+)', id_number)
        if match:
            prefix_alpha = match.group(1)
            range_start = int(match.group(2))
            range_end = int(match.group(3))
            if range_end < range_start:
                error_statement = f"Error: End of range ({range_end}) is less than start of range ({range_start}) for ID number {id_number}"
                new_errors.append(error_statement)
                #new_id_numbers.append(f"{prefix_alpha}{range_start}")
            else:
                for i in range(range_start, range_end + 1):
                    new_id_numbers.append(f"{prefix_alpha}{i}")
        else:
            # Check if the second pattern matches
            match = re.match(r'([A-Za-z]+\d+/?\d*)/(\d+)-(\d+)', id_number)
            if match:
                prefix_alpha = match.group(1)
                prefix_numeric = match.group(2)
                range_start = int(prefix_numeric)
                range_end = int(match.group(3))
                if range_end < range_start:
                    #new_id_numbers.append(f"{prefix_alpha}{range_start}")
                    error_statement = f"Error: End of range ({range_end}) is less than start of range ({range_start}) for ID number {id_number}"
                    new_errors.append(error_statement)
                else:
                    # Generate new id numbers based on the matched pattern
                    for i in range(range_start, range_end + 1):
                        new_id_numbers.append(f"{prefix_alpha}/{i:02}")
            else:
                # If neither pattern matches, append the id_number as is
                new_id_numbers.append(id_number)
    return new_id_numbers


def get_manufacture_model(description: str):
    workbook = load_workbook("../database/Full_list_of_Manufacturers_and_Models.xlsx")
    sheet = workbook['Model']

    description_keywords = description.lower()
    manufacture, model = "",""

    # Search for Model in the column
    for row in sheet.iter_rows(min_row=2, values_only=True):
        model_keyword = row[0]

        if model_keyword and str(model_keyword).lower() in description_keywords:
            model = model_keyword
            manufacture = row[1]
            break

    return manufacture, model


def contains_keyword(first_row, keyword):
    if first_row is not None:
        row_text = ' '.join(cell if cell is not None else '' for cell in first_row)
        if row_text.lower().startswith(keyword.lower()):
            return True
    return False


def add_six_months(date_str):

    date_declaration_obj = datetime.strptime(date_str, '%d/%m/%Y')
    # Add 6 months to the date of declaration
    next_inspection_date_obj = date_declaration_obj + timedelta(days=6*30)
    # Convert the next inspection date back to a string in the same format
    next_inspection_date_str = next_inspection_date_obj.strftime('%d/%m/%Y')

    return next_inspection_date_str

# Call to the First Integrated PDF
def extract_first_integrated_pdf(pdf_path):
    print("<------------extracting first_integrated pdf------------>")
    pdf_doc = pdfplumber.open(pdf_path)
    extraction_info = dict()
    page_errors = dict()
    for i in range(0, len(pdf_doc.pages)):
        try:
            # for page_number, page in enumerate(pdf_doc.pages()):
            page = pdf_doc.pages[i]
            page_tables = page.extract_tables()
            #print("page number:", i)
            #print("page tables:", page_tables)
            if not page_tables:
                page_errors[i+1] = f" No tables found on page {i+1}. Skipping..."
                continue

            first_row = page_tables[0][0]

            if contains_keyword(first_row, "Name & Address of employer for Whom the examination was made"):
                process_table_type1(page_tables, extraction_info)
            elif contains_keyword(first_row, "Date of Thorough Examination"):
                process_table_type2(page_tables[0], extraction_info)
            elif contains_keyword(first_row, "Name &AddressofManufacturer") or contains_keyword(first_row, "Name & Address of Manufacturer"):
                process_table_type3(page_tables[0], extraction_info)
            else:
                page_errors[i+1] = f"No recognized table found on page {i+1}"

        except Exception as e:
            page_errors[i+1] = f"Error occurred on page {i+1}: {e}"
            print(f"Error occurred on page {i+1}: {e}")

    excel_management.create_excel(extraction_info, "../database/First Integrated.xlsx", "First_Integrated", page_errors)


def process_table_type1(page_tables, extraction_info):
    try:
        id_number, description = None, None
        id_pattern = re.compile(r'(\(\w+\)\s*)?[A-Z]{3}\d+')
        swl_pattern = re.compile(r'(\d+\.\d+)\s+(Tonnes|Kilos)\s', re.IGNORECASE)
        row_data = dict()
        page_errors = dict()
        errors = list()
        report_ref_no, date_of_thorough_examination = None, None
        for page_num, page in enumerate(page_tables):
            try:
                for index, row in enumerate(page):
                    page_info = {}
                    row_text = ' '.join(cell if cell is not None else '' for cell in row)

                    if "Name & Address of employer" in row_text:
                        continue        # Skip this row
                        # Extract ID Number
                    id_match = id_pattern.search(row_text)
                    if id_match:
                        try:
                            id_number = id_match.group()
                            if not id_number:
                                errors.append("ID Number not found")
                        except Exception as e:
                            errors.append("Error extracting ID Number: {e}")
                        #print("id_number",id_number)
                        #page_info["Id Number"] = id_number

                        # Extract SWL
                        swl_match = swl_pattern.search(row_text)
                        if swl_match:
                            try:
                                swl_value = swl_match.group(1)
                                swl_units = swl_match.group(2)
                                swl_note = None
                                page_info["SWL Value"] = swl_value
                                page_info["SWL Unit"] = swl_units
                                page_info["SWL Note"] = swl_note
                            except Exception as e:
                                errors.append("Error extracting SWL: {e}")

                        # Extract Description between ID Number and SWL
                        if id_match:
                            start_index = id_match.end()
                            if swl_match:
                                end_index = swl_match.start()
                            else:
                                end_index = -1
                            try:
                                description = row_text[start_index:end_index].strip()
                                description = re.sub(r'\([^)]*\)\s*', '', description)
                                if not description:
                                    errors.append("Description not found")
                                else:
                                    page_info["Item Description"] = description
                            except Exception as e:
                                errors.append("Error extracting Description: {e}")
                            #print("description", description)

                            #Get manufacture and model fro  description
                            try:
                                manufacture, model = get_manufacture_model(description)
                                if not manufacture:
                                    errors.append("Manufacturer not found")
                                else:
                                    page_info["Manufacturer"] = manufacture
                                if not model:
                                    errors.append("Model not found")
                                else:
                                    page_info["Model"] = model
                            except Exception as e:
                                errors.append("Error extracting Manufacture and Model: {e}")

                        # Extract Date of Next Inspection
                        next_inspection_pattern = re.compile(r'(\d{2}/\d{2}/\d{4})$', re.IGNORECASE)
                        next_inspection_match = next_inspection_pattern.search(row_text)
                        if next_inspection_match:
                            try:
                                next_inspection_date = next_inspection_match.group(1).strip()
                                if not next_inspection_date:
                                    errors.append("Next Inspection date not found")
                                else:
                                    page_info["Next Inspection Due Date"] = next_inspection_date
                            except Exception as e:
                                errors.append("Error extracting Next Inspection date: {e}")
                            #print("Next Inspection Date:", next_inspection_match)

                        row_data[id_number] = page_info

            except Exception as e:
                page_errors[page_num] = "Error", e, " occurred while processing the page:"
                print("Error", e, " occurred while processing the page:", page_num)

        # Extract Date of Previous Inspection and Certificate Number
        for page_num, page in enumerate(page_tables):
            try:
                for row in page:
                    row_text = ' '.join(cell if cell is not None else '' for cell in row)
                    # Extract Date of Previous Inspection
                    keyword_date = "Date of thorough examination"
                    try:
                        if keyword_date.lower() in row_text.lower():
                            date_pattern = re.compile(r'\b(\d{2}/\d{2}/\d{4})\b')
                            previous_inspection_match = date_pattern.search(row_text)
                            if previous_inspection_match:
                                date_of_thorough_examination = previous_inspection_match.group().strip()
                                if not date_of_thorough_examination:
                                    errors.append("Previous Inspection date not found")
                    except Exception as e:
                        errors.append("Error extracting Previous Inspection date: {e}")

                    # Extract Report No
                    keyword_report = "Report Ref No"
                    try:
                        if keyword_report.lower() in row_text.lower():
                            report_pattern = re.compile(r'[A-Z]{3}/\d{6}/\d{5}')
                            report_ref_match = report_pattern.search(row_text)
                            if report_ref_match:
                                report_ref_no = report_ref_match.group().strip()
                                if not report_ref_no:
                                    errors.append("Certificate Number not found")
                    except Exception as e:
                        errors.append("Error extracting Report Ref: {e}")

                # Assign Previous Inspection Date and Certificate No to all items in row_data
                for id_number, page_info in row_data.items():
                    page_info["Previous Inspection"] = date_of_thorough_examination
                    page_info["Certificate No"] = report_ref_no
                    extraction_info[id_number] = page_info

            except Exception as e:
                page_errors[page_num] = "Error", e, " occurred while processing the page:"
                print("Error", e, " occurred while processing the page:", page_num)

    except Exception as e:
        print("Error occurred in process_table_type1:", e)


def process_table_type2(table, extraction_info):
    try:
        page_info = {}
        page_errors = dict()
        errors = list()
        id_number, description = None, None
        report_number_pattern = re.compile(r'Report\s*Number:', re.IGNORECASE)
        for page_num, row_outer in enumerate(table):
            try:
                #print("row_outer", row_outer)
                for cell in row_outer:
                    if cell:
                        # Extract information for id number, description, swl
                        if "identification of the equipment" in cell:
                            try:
                                parts = cell.split('\n')
                                if len(parts) >= 2:
                                    description = parts[1].strip()
                                    id_number = parts[2].strip()
                                    if not description:
                                        errors.append("Description not found")
                                    else:
                                        page_info["Item Description"] = description
                                    if not id_number:
                                        errors.append("ID Number not found")
                                    else:
                                        id_number_list = []
                                        if "-" in id_number:
                                            split_id_numbers = split_id_numbers_with_range([id_number])
                                            print("new_id_numbers", split_id_numbers)
                                            id_number_list.extend(split_id_numbers)
                                            print("id_number_list", id_number_list)
                                        else:
                                            id_number_list.append(id_number)
                                        for id_number in id_number_list:
                                            #print("id_number_list", id_number)
                                            extraction_info[id_number] = page_info
                            except Exception as e:
                                errors.append(f"Error extracting ID Number or Description: {e}")
                            #Get manufacture and model from description
                            try:
                                manufacture, model = get_manufacture_model(description)
                                if not manufacture:
                                    errors.append("Manufacturer not found")
                                else:
                                    page_info["Manufacturer"] = manufacture
                                if not model:
                                    errors.append("Model not found")
                                else:
                                    page_info["Model"] = model
                            except Exception as e:
                                errors.append(f"Error extracting manufacture and model: {e}")

                        elif "WLL" in cell:
                            try:
                                parts = cell.split('\n')
                                if len(parts) >= 4:
                                    wll = parts[3].strip()
                                    swl_pattern = re.compile(r'(\d+(?:\.\d+)?)\s*(TONNE|TONNES)\b', re.IGNORECASE)
                                    swl_match = swl_pattern.search(wll)
                                    if swl_match:
                                        swl_value = swl_match.group(1)
                                        swl_units = swl_match.group(2)
                                        swl_note = None
                                        page_info["SWL Value"] = swl_value
                                        page_info["SWL Unit"] = swl_units
                                        page_info["SWL Note"] = swl_note
                            except Exception as e:
                                errors.append(f"Error extracting SWL: {e}")

                        elif report_number_pattern.search(cell):
                            try:
                                certificate_number = report_number_pattern.sub('', cell).strip()
                                if not certificate_number:
                                    errors.append("Certificate Number not found")
                                else:
                                    page_info["Certificate No"] = certificate_number
                            except Exception as e:
                                errors.append(f"Error extracting Certificate Number: {e}")
                        elif "Date of Thorough" in cell:
                            try:
                                previous_inspection = cell.replace("Date of Thorough Examination:", "").strip()
                                if not previous_inspection:
                                    errors.append("Previous Inspection not found")
                                else:
                                    page_info["Previous Inspection"] = previous_inspection
                            except Exception as e:
                                errors.append(f"Error extracting Previous Inspection: {e}")
                        elif "Latest date by which next" in cell:
                            try:
                                next_inspection = cell.replace(
                                    "Latest date by which next thorough\nexamination must be carried out:", "").strip()
                                if not next_inspection:
                                    errors.append("Next Inspection Due Date not found")
                                else:
                                    page_info["Next Inspection Due Date"] = next_inspection
                            except Exception as e:
                                errors.append(f"Error extracting Next Inspection: {e}")

            except Exception as e:
                page_errors[page_num] = "Error", e, " occurred while processing the page:"
                print("Error", e, " occurred while processing the page:", page_num)

            if errors:
                errors.append("page no: " + str(page_num + 1))
                page_info["Errors"] = str(errors)

    except Exception as e:
        print("Error occurred in process_table_type2:", e)


def process_table_type3(table, extraction_info):
    try:
        page_info = dict()
        page_errors = dict()
        errors = list()
        id_number, description, date_of_declaration = None, None, None
        for i, row in enumerate(table):
            try:
                if row:
                    # Check for labels
                    for j, cell in enumerate(row):
                        if cell:
                            # Check for Id Number label
                            if "Id Number" in cell:
                                try:
                                    # Get the value from the next row
                                    id_number_row = table[i + 1] if i + 1 < len(table) else None
                                    if id_number_row:
                                        id_number = id_number_row[j].strip() if j < len(id_number_row) else None
                                        id_number_list = []
                                        if "-" in id_number:
                                            split_id_numbers = split_id_numbers_with_range([id_number])
                                            id_number_list.extend(split_id_numbers)
                                        else:
                                            id_number_list.append(id_number)
                                        for id_number in id_number_list:
                                            extraction_info[id_number] = page_info
                                    elif not id_number_row:
                                        errors.append("Id Number not found")
                                except Exception as e:
                                    errors.append(f"Error extracting Id Number: {e}")

                            # Check for Description label
                            elif "Description" in cell:
                                try:
                                    parts = cell.split('\n')
                                    if len(parts) >= 2:
                                        description = parts[1].strip()
                                        page_info["Item Description"] = description
                                    elif not description:
                                        errors.append("Description not found")
                                except Exception as e:
                                    errors.append(f"Error extracting Description: {e}")

                            # Get manufacture and model from description
                                try:
                                    manufacture, model = get_manufacture_model(description)
                                    if not manufacture:
                                        errors.append("Manufacturer not found")
                                    else:
                                        page_info["Manufacturer"] = manufacture
                                    if not model:
                                        errors.append("Model not found")
                                    else:
                                        page_info["Model"] = model
                                except Exception as e:
                                    errors.append(f"Error extracting manufacture and model: {e}")
                            # Check for WLL label
                            elif "WLL" in cell:
                                try:
                                    # Get the value from the next row
                                    wll_row = table[i + 1] if i + 1 < len(table) else None
                                    if wll_row:
                                        wll = wll_row[j].strip() if j < len(wll_row) else None
                                        swl_pattern = re.compile(r'(\d+(?:\.\d+)?)\s*(TONNE|TONNES)\b', re.IGNORECASE)
                                        swl_match = swl_pattern.search(wll)
                                        if swl_match:
                                            swl_value = swl_match.group(1)
                                            swl_units = swl_match.group(2)
                                            page_info["SWL Value"] = swl_value
                                            page_info["SWL Unit"] = swl_units
                                except Exception as e:
                                    errors.append(f"Error extracting SWL: {e}")
                            elif "Certificate Number" in cell:
                                try:
                                    certificate_row = table[i + 1] if i + 1 < len(table) else None
                                    if certificate_row:
                                        certificate_number = certificate_row[j].strip() if j < len(
                                            certificate_row) else None
                                        page_info["Certificate No"] = certificate_number
                                    elif not certificate_row:
                                        errors.append("Certificate Number not found")
                                except Exception as e:
                                    errors.append(f"Error extracting Certificate Number: {e}")
                            elif "Date of Declaration" in cell:
                                try:
                                    parts = cell.split('\n')
                                    if len(parts) >= 3:
                                        date_of_declaration = parts[3].strip()
                                        # Use the function to add 6 months to the date
                                        next_inspection_date = add_six_months(date_of_declaration)
                                        page_info["Previous Inspection"] = date_of_declaration
                                        page_info["Next Inspection Due Date"] = next_inspection_date
                                    elif not date_of_declaration:
                                        errors.append("Date of Declaration not found")

                                except Exception as e:
                                    errors.append(f"Error extracting Date of Declaration: {e}")

            except Exception as e:
                page_errors[i] = "Error", e, " occurred while processing the page:"
                print("Error", e, " occurred while processing the page:", i)

            if errors:
                errors.append("page no: " + str(i + 1))
                page_info["Errors"] = str(errors)

    except Exception as e:
        print("Error occurred in process_table_type3:", e)

if __name__ == "__main__":
    extract_first_integrated_pdf("../resources/First Integrated Full Cert Pack.pdf")
