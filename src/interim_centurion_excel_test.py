import openpyxl


column_map = dict()
column_map["Certificate Number"] = "A"
column_map["Description"] = "D"
column_map["WWL"] = "F"
column_map["Previous Examination"] = "K"
column_map["Ref No"] = "Q"
column_map["Next Examination"] = "N"
column_map["Manufacturer"] = "G"
column_map["Model"] = "I"


def find_last_row(sheet):
    # last_row = sheet.max_row
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        if row[0].value is None:
            return row[0].row - 1  # Return the row before the first empty cell
    return sheet.max_row


def update_excel(extracted_data: dict, sheet_name: str):
    print("<--------------Updating excel------------------>")
    workbook = openpyxl.load_workbook("../database/Centurion_data.xlsx")
    # Select the specified sheet
    sheet = workbook[sheet_name]
    # Find the last row with data in column A
    last_row = find_last_row(sheet) + 1
    for key, data in extracted_data.items():
        sheet[f"A{last_row}"] = key
        for cell_name, value in data.items():
            # Create cell address based on column name and current row
            column_name = column_map[cell_name]
            cell_address = f"{column_name}{last_row}"
            sheet[cell_address] = value
        last_row += 1
    workbook.save("../database/Centurion_data.xlsx")
    # Close the workbook
    workbook.close()
    print("<-------------- excel updated successfully------------------>")


if __name__ == "__main__":
    update_excel(dict())