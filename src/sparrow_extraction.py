import re
import pdfplumber
from openpyxl import load_workbook
import excel_management


def get_manufacture_model(description: str):
    workbook = load_workbook("database/Full_list_of_Manufacturers_and_Models.xlsx")
    manufacturer_sheet = workbook['Manufacture']
    model_sheet = workbook['Model']
    description_keywords = description.lower().split()
    manufacture, model = "", ""
    for row in manufacturer_sheet.iter_rows(min_row=2, values_only=True):
        keyword = row[0]
        value = row[1]
        if keyword and str(keyword).lower() in description_keywords:
            manufacture = value
            break
    if str(manufacture).lower() == "miller" and "weblift" in description_keywords:
        manufacture = "Miller Weblift"
    for row in model_sheet.iter_rows(min_row=2, values_only=True):
        keyword = row[0]
        value = row[1]
        if keyword and str(keyword).lower() in description_keywords:
            model = keyword
            if not manufacture and value:
                manufacture = value
            break
    return manufacture, model


def process_swl(swl: str):
    pattern = r'^(\d+(?:\.\d+)?)([a-zA-Z]+)?\s*(.*)$'

    # Match the pattern
    match = re.match(pattern, swl)

    if match:
        value_part = match.group(1)
        unit_part = match.group(2)
        note_part = match.group(3)
    else:
        value_part = None
        unit_part = None
        note_part = swl
    units_map = {"kgs": "kg"}
    if unit_part in units_map:
        unit_part = units_map[unit_part]
    # Check if part 2 is a unit type or not
    units = ["kg", "g", "lb", "ton", "t", "m", "cm", "mm", "ft", "in", "m²", "cm²", "mm²", "ft²", "in²", "m³", "cm³", "mm³",
             "ft³", "in³", "km/h", "mph", "m/s", "kph", "°C", "°F", "°K", "bar", "atm", "Pa", "kPa", "psi", "N", "J",
             "W", "A", "V", "F", "Ω", "S", "H", "Hz", "C", "Bq", "Gy", "Sv", "cd", "lm", "lx", "B", "mol", "unit", "te"]
    if unit_part and unit_part not in units:
        note_part = unit_part + " " + note_part if note_part else unit_part
        unit_part = None

    return value_part, unit_part, note_part


def get_identification_parts_list(input_string: str, quantity: int):
    numeric_part = ''.join(filter(str.isdigit, input_string))
    part_list = list()
    if input_string[0].isdigit():
        alpha_part = input_string[len(numeric_part):]
        for i in range(0, quantity):
            part_list.append(f"{int(numeric_part) + i}{alpha_part}")
    else:
        alpha_part = input_string[:len(input_string)-len(numeric_part)]
        for i in range(0, quantity):
            part_list.append(f"{alpha_part}{int(numeric_part) + i}")
    return part_list


def get_identification_number_list(identification_numbers: str, quantity: int):
    #Take this as example (D971-1 to 6) or (MGL1 to MGL36)
    if "to" in identification_numbers.lower():
        # identification_number_first_part = D971-1 or MGL1
        identification_number_first_part = identification_numbers.split("to")[0].strip()
        # print(identification_number_first_part)
        # example: D971-1
        if "-" in identification_number_first_part:
            # first_part = D971, second_part = 1
            first_part, second_part = identification_number_first_part.split('-')
            # print(first_part, second_part)
            second_part_list = get_identification_parts_list(second_part, quantity)
            # print(second_part_list)
            identification_number_list = list()
            for second_part in second_part_list:
                identification_number_list.append(f"{first_part}-{second_part}")
        else:
            # example: MGL1
            identification_number_list = get_identification_parts_list(identification_number_first_part, quantity)
    elif re.search(r'x(\d+)', identification_numbers):
        # print(identification_numbers)
        identification_number_list = list()
        for num in identification_numbers.split(','):
            id_number = num.split('x')[0].strip()
            for i in range(len(id_number)-1, -1, -1):
                if id_number[i].isdigit():
                    id_number= id_number[:i+1]
                    break
            count = int(''.join(filter(str.isdigit, num.split('x')[1].strip())))
            for i in range(1,count+1):
                identification_number_list.append(f"{id_number}-{i}")
    elif "," in identification_numbers:
        identification_number_list = identification_numbers.split(',')

    # print(identification_number_list)
    return identification_number_list


def extract_sparrow_pdf(pdf_path):
    try:
        print("<------------extracting sparrow pdf------------>")
        pdf_doc = pdfplumber.open(pdf_path)
        extraction_info = dict()
        page_errors = dict()
        for i in range(0, len(pdf_doc.pages)):
            try:
                page = pdf_doc.pages[i]
                table_extract = page.extract_tables()
                if table_extract:
                    print("page number:", i+1)
                    page_tables = table_extract[0]
                    table_data1 = page_tables[0][0].split('\n')
                    table_data3 = page_tables[3]
                    table_data4 = page_tables[4]
                    identification_numbers = description = swl = quantity = None
                    errors = list()
                    for index in range(0, len(table_data3)):
                        try:
                            if table_data3[index] is None:
                                continue
                            text_to_compare = table_data3[index].lower()
                            if not identification_numbers and "identification" in text_to_compare:
                                identification_numbers = table_data4[index].strip()
                            elif not description and "description" in text_to_compare:
                                description = table_data4[index].replace('\n', ' ')
                            elif not swl and "swl" in text_to_compare:
                                swl = table_data4[index].strip()
                            elif not quantity and "quantity" in text_to_compare:
                                quantity = int(float(table_data4[index]))
                        except Exception as e:
                            print("Error extracting value from page:", e)

                    if identification_numbers:
                        page_info = dict()
                        # page_info["Id Number"] = table_data4[0].strip()
                        if description:
                            try:
                                item_description = description
                                if not item_description:
                                    errors.append("Item Description not found")
                                else:
                                    page_info["Item Description"] = item_description
                            except Exception as e:
                                errors.append(e)
                            try:
                                manufacturer, model = get_manufacture_model(description)
                                if not manufacturer:
                                    errors.append("Manufacturer not found")
                                else:
                                    page_info["Manufacturer"] = manufacturer
                                if not model:
                                    errors.append("Model not found")
                                else:
                                    page_info["Model"] = model
                            except Exception as e:
                                errors.append(e)
                        else:
                            errors.append(
                                "Description not found in the page. Item Description, Manufacturer, Model columns are left empty")
                        # page_info["SWL"] = swl
                        if swl:
                            try:
                                swl_value, swl_unit, swl_note = process_swl(swl)
                                if not swl_value:
                                    errors.append("SWL Value not found")
                                else:
                                    page_info[
                                        "SWL Value"] = swl_value
                                if not swl_unit:
                                    errors.append("SWL Unit not found")
                                else:
                                    page_info["SWL Unit"] = swl_unit
                                page_info["SWL Note"] = swl_note
                            except Exception as e:
                                errors.append(e)
                        else:
                            errors.append(
                                "SWL not found in the page")
                        # report_number, date_of_examination, job_number, next_date_of__examination = None, None, None, None

                        table_data1_mapping = dict()
                        for data in table_data1:
                            try:
                                data_list = data.split(':')
                                key = data_list[0].lower().replace(" ", "").strip()
                                value = data_list[-1].strip()
                                table_data1_mapping[key] = value
                            except Exception as e:
                                print("Error extracting value from page:", e)

                        if "reportnumber" not in table_data1_mapping:
                            errors.append("Certificate no not found")
                        else:
                            page_info["Certificate No"] = table_data1_mapping["reportnumber"]
                        if "dateofthoroughexamination" not in table_data1_mapping:
                            errors.append("Previous Inspection not found")
                        else:
                            page_info["Previous Inspection"] = table_data1_mapping["dateofthoroughexamination"]
                        if "jobnumber" not in table_data1_mapping:
                            errors.append("Provider Identification not found")
                        else:
                            page_info["Provider Identification"] = "LOFT-" + table_data1_mapping["jobnumber"]
                        if "duedateofnextthoroughexamination" not in table_data1_mapping:
                            errors.append("Next Inspection Due Date not found")
                        else:
                            page_info["Next Inspection Due Date"] = table_data1_mapping["duedateofnextthoroughexamination"]

                        try:
                            if quantity > 1:
                                identification_number_list = get_identification_number_list(identification_numbers,
                                                                                            quantity)
                            else:
                                identification_number_list = list()
                                identification_number_list.append(identification_numbers)
                            if errors:
                                errors.append("page no: "+str(i+1))
                                page_info["Errors"] = str(errors)
                                # print(identification_numbers, errors)
                            for identification_number in identification_number_list:
                                extraction_info[identification_number] = page_info
                        except Exception as e:
                            errors.append(
                                "Error in extracting identification numbers. So, appending the identification number as found in the page")
                            errors.append("page no: " +str(i+1))
                            print("Error in extracting identification numbers. So, appending the identification number as found in the page")
                            page_info["Errors"] = str(errors)
                            extraction_info[identification_numbers] = page_info
                        # print(identification_numbers, page_info)
                    else:
                        page_errors[i+1] = "No identification numbers are found in the page. So, the page is not processed."
                        print("No identification number found")
                else:
                    page_errors[i+1] = "No text found on page. probably it's an image. So, the page is not processed."
            except Exception as e:
                page_errors[i+1] = "Error" + str(e) + " occurred while processing the page:"
                print("Error", {e}, " occurred while processing the page:", i)

        print(len(extraction_info.keys()), page_errors.keys())
        excel_management.create_excel(extraction_info, "database/Sparrows.xlsx", "Sparrows", page_errors)
    except Exception as e:
        print("An error occurred:", e)


if __name__ == "__main__":
    extract_sparrow_pdf("resources/sparrows.pdf")
    # get_identification_number_list("A1 to A6", 8)
