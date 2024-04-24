import re
import pdfplumber
from openpyxl import load_workbook
from datetime import datetime

import excel_management

def get_manufacture_model(description: str):
    workbook = load_workbook("../database/Full_list_of_Manufacturers_and_Models.xlsx")
    manufacturer_sheet = workbook['Manufacture']
    model_sheet = workbook['Model']
    description_keywords = description.lower().split()
    manufacture, model = "", ""
    for row in manufacturer_sheet.iter_rows(min_row=2, values_only=True):
        keyword = row[0]
        value = row[1]
        if keyword and str(keyword).lower() in description_keywords:
            manufacture = value
            break
    # if str(manufacture).lower() == "miller" and "weblift" in description_keywords:
    #     manufacture = "Miller Weblift"
    for row in model_sheet.iter_rows(min_row=2, values_only=True):
        keyword = row[0]
        value = row[1]
        if keyword and str(keyword).lower() in description_keywords:
            model = keyword
            if not manufacture and value:
                manufacture = value
            break
    return manufacture, model


def process_swl(swl: str):
    pattern = r'^(\d+(?:\.\d+)?)([a-zA-Z]+)?\s*(.*)$'

    # Match the pattern
    match = re.match(pattern, swl)

    if match:
        value_part = match.group(1)
        unit_part = match.group(2)
        note_part = match.group(3)
    else:
        value_part = None
        unit_part = None
        note_part = swl
    units_map = {"kgs": "kg"}
    if unit_part in units_map:
        unit_part = units_map[unit_part]
    # Check if part 2 is a unit type or not
    units = ["kg", "g", "lb", "ton", "t", "m", "cm", "mm", "ft", "in", "m²", "cm²", "mm²", "ft²", "in²", "m³", "cm³", "mm³",
             "ft³", "in³", "km/h", "mph", "m/s", "kph", "°C", "°F", "°K", "bar", "atm", "Pa", "kPa", "psi", "N", "J",
             "W", "A", "V", "F", "Ω", "S", "H", "Hz", "C", "Bq", "Gy", "Sv", "cd", "lm", "lx", "B", "mol", "unit", "te", "KG", "T", "TE", "G", "TH /FITTING"]
    if unit_part and unit_part not in units:
        note_part = unit_part + " " + note_part if note_part else unit_part
        unit_part = None

    return value_part, unit_part, note_part


def get_identification_parts_list(input_string: str, quantity: int):
    numeric_part = ''.join(filter(str.isdigit, input_string))
    part_list = list()
    if input_string[0].isdigit():
        alpha_part = input_string[len(numeric_part):]
        for i in range(0, quantity):
            part_list.append(f"{int(numeric_part) + i}{alpha_part}")
    else:
        alpha_part = input_string[:len(input_string)-len(numeric_part)]
        for i in range(0, quantity):
            part_list.append(f"{alpha_part}{int(numeric_part) + i}")
    return part_list


def get_identification_number_list(identification_numbers: str, quantity: int):
    identification_number_list = []
    if "to" in identification_numbers.lower():
        # identification_number_first_part = D971-1 or MGL1
        identification_number_first_part = identification_numbers.split("to")[0].strip()
        # print(identification_number_first_part)
        # example: D971-1
        if "-" in identification_number_first_part:
            # first_part = D971, second_part = 1
            first_part, second_part = identification_number_first_part.split('-')
            # print(first_part, second_part)
            second_part_list = get_identification_parts_list(second_part, quantity)
            # print(second_part_list)
            identification_number_list = list()
            for second_part in second_part_list:
                identification_number_list.append(f"{first_part}-{second_part}")
        else:
            # example: MGL1
            identification_number_list = get_identification_parts_list(identification_number_first_part, quantity)
    elif re.search(r'x(\d+)', identification_numbers):
        # print(identification_numbers)
        identification_number_list = list()
        for num in identification_numbers.split(','):
            id_number = num.split('x')[0].strip()
            for i in range(len(id_number)-1, -1, -1):
                if id_number[i].isdigit():
                    id_number= id_number[:i+1]
                    break
            count = int(''.join(filter(str.isdigit, num.split('x')[1].strip())))
            for i in range(1,count+1):
                identification_number_list.append(f"{id_number}-{i}")
    elif "-" in identification_numbers and not identification_numbers.replace('-', '').isdigit():

        parts = identification_numbers.split("-")
        if len(parts) == 2:
            prefix_part = parts[0][:-1]
            start_number = int(parts[0][-1])
            end_number = int(parts[1])
            identification_number_list = [
                f"{prefix_part}{i}" for i in range(start_number, end_number + 1)
            ]
        else:
            identification_number_list.append(identification_numbers.strip())

    return identification_number_list



def extract_quantity(text):
    match = re.match(r'^\s*(\d+)\s+\d*', text)
    if match:
        return int(match.group(1))
    else:
        return None


def extrac_serialnumber(cell_content):
    serial_numbers_raw = cell_content.split(',')
    serial_numbers_cleaned = []
    for serial in serial_numbers_raw:
        cleaned_serial = serial.strip()
        match = re.search(r'(.+?)\s+-', cleaned_serial)
        if match:
            cleaned_serial = match.group(1).strip()
        else:
            cleaned_serial = cleaned_serial
        cleaned_serial = cleaned_serial.replace(" ", "")
        prefix = "SerialNo(s):"
        if cleaned_serial.startswith(prefix):
            cleaned_serial = cleaned_serial[len(prefix):]

        serial_numbers_cleaned.append(cleaned_serial)
        cleaned_serials_string = ','.join(serial_numbers_cleaned)
    return cleaned_serials_string


def extraction_centurion_pdf(pdf_path):
    print("<------------extracting centurion pdf------------>")
    pdf = pdfplumber.open(pdf_path)
    extraction_info = dict()
    page_errors = dict()
    for i, page in enumerate(pdf.pages):
        text = page.extract_text()
        page = pdf.pages[i]
        if "Centurion" in text and "Hendrik" not in text:
            if page.extract_tables():
                try:
                    print("page number:", i)
                    page_tables = page.extract_tables()
                    first_table = page_tables[0]
                    # data1:Report Number / Date of Examination / Ref No
                    table_data1 = first_table[0][13]
                    # data2: Identify Company keywords
                    table_data2 = first_table[1]
                    # data3: Identify Text keywords
                    table_data3 = first_table[2]
                    # data4: Provide related Value
                    table_data4 = first_table[4]
                    # data5 ID number Value
                    table_data5 = first_table[5]
                    full_serials = ' '.join([item for item in table_data5 if item is not None])
                    errors = list()
                    identification_number_list = list()
                except Exception as e:
                    print("Error extracting format from page:", e)
                    continue
                if "Quantity & Description of Equipment, Serial Numbers" in table_data3[0]:

                    id_numbers, description, mnfer, wwl, next_thorough = None, None, None, None, None
                    quantity = 1

                    for index in range(0, len(table_data3)):
                        try:
                            if table_data3[index] is None:
                                continue
                            text_to_compare = table_data3[index].lower()
                            if not description and "description" in text_to_compare:
                                description = table_data4[0].replace('\n', ' ')
                                serial_numbers = full_serials
                                serial = re.sub(r'\s+', '', serial_numbers)
                                serial_cleaned = extrac_serialnumber(serial)
                                mnfer = table_data4[4].strip()

                            elif not wwl and "working" in text_to_compare:
                                wwl = table_data4[index].strip()
                            elif not next_thorough and "next" in text_to_compare:
                                date_string = table_data4[index].strip()
                                date_obj = datetime.strptime(date_string, "%d/%m/%Y")
                                next_thorough = date_obj.strftime("%d/%m/%Y")
                            elif not id_numbers and "certificate" in text_to_compare:
                                id_numbers = table_data4[index].strip()
                        except Exception as e:
                            print("Error extracting value from page:", e)


                    if id_numbers:
                        page_info = dict()
                        if description:
                            try:
                                item_description = description
                                if not item_description:
                                    errors.append("Item Description not found")
                                else:
                                    page_info["Item Description"] = description.split(':')[0]
                            except Exception as e:
                                errors.append(e)
                            try:
                                manufacturer, model = get_manufacture_model(description)
                                manufacturer = mnfer
                                if not manufacturer:
                                    errors.append("Manufacturer not found")
                                else:
                                    page_info["Manufacturer"] = manufacturer
                                if not model:
                                    errors.append("Model not found")
                                else:
                                    page_info["Model"] = model
                            except Exception as e:
                                errors.append(e)

                        if wwl:
                            try:
                                swl_value, swl_unit, swl_note = process_swl(wwl)
                                if not swl_value:
                                    errors.append("SWL Value not found")
                                else:
                                    page_info["SWL Value"] =swl_value
                                if not swl_unit:
                                    errors.append("SWL Unit not Found")
                                else:
                                    page_info["SWL Unit"] = swl_unit
                                page_info["SWL Note"] = swl_note
                            except Exception as e:
                                errors.append(e)
                        else:
                            errors.append("SWL not found in this page.")
                        page_info["Next Inspection Due Date"] = next_thorough
                        # report_number, date_of_examination, job_number, next_date_of__examination = None, None, None, None
                        table_data1_mapping = dict()
                        table_data1 = table_data1.splitlines()

                        for data in table_data1:
                            data_list = data.split(':', 1)
                            if len(data_list) == 2:
                                key, value = data_list
                                formattted_key = key.lower().replace(" ", "").replace("/", "").replace(".", "")
                                table_data1_mapping[formattted_key] = value.strip()

                        page_info["Provider Identification"] = table_data1_mapping["custrefpono"]
                        page_info["Certificate No"] = table_data1_mapping["reportnumber"]
                        page_info["Previous Inspection"] = table_data1_mapping["dateofexamination"]

                        id_numbers = serial_cleaned

                        if quantity == 1:
                            identification_number_list.append(id_numbers)

                        for identification_number in identification_number_list:
                            extraction_info[identification_number] = page_info

                    else:
                        print("No identification error")

                elif "Qty, Description of Equipment, Serial Numbers" in table_data3[0]:
                    quantity, id_numbers, description, mnfer, wwl, next_thorough = None, None, None, None, None, None
                    for index in range(0, len(table_data3)):
                        if table_data3[index] is None:
                            continue
                        text_to_compare = table_data3[index].lower()
                        if not description and "description" in text_to_compare:
                            description = table_data4[index].replace('\n', ' ')
                            serial_numbers = full_serials
                            serial = re.sub(r'\s+', '', serial_numbers)
                            serial_cleaned = extrac_serialnumber(serial)
                            quantity = extract_quantity(table_data4[index])
                            mnfer = table_data4[4].strip()
                        elif not wwl and "working" in text_to_compare:
                            wwl = table_data4[index].strip()
                        elif not next_thorough and "next" in text_to_compare:
                            date_string = table_data4[index].strip()
                            date_obj = datetime.strptime(date_string, "%d/%m/%Y")
                            next_thorough = date_obj.strftime("%d/%m/%Y")
                        elif not id_numbers and "certificate" in text_to_compare:
                            id_numbers = table_data4[index].strip()
                            old_id = id_numbers
                            id_numbers = serial_cleaned


                    if id_numbers:
                        page_info = dict()
                        if description:
                            page_info["Item Description"] = description.split(':')[0]
                            manufacturer, model = get_manufacture_model(description)
                            manufacturer = mnfer
                            page_info["Manufacturer"] = manufacturer
                            page_info["Model"] = model
                        if wwl:
                            try:
                                swl_value, swl_unit, swl_note = process_swl(wwl)
                                if not swl_value:
                                    errors.append("SWL Value not found")
                                else:
                                    page_info["SWL Value"] = swl_value
                                if not swl_unit:
                                    errors.append("SWL Unit not Found")
                                else:
                                    page_info["SWL Unit"] = swl_unit
                                page_info["SWL Note"] = swl_note
                            except Exception as e:
                                errors.append(e)
                        else:
                            errors.append("SWL not found in this page.")
                        page_info["Next Inspection Due Date"] = next_thorough
                        # report_number, date_of_examination, job_number, next_date_of__examination = None, None, None, None
                        table_data1_mapping = dict()
                        table_data1 = table_data1.splitlines()

                        for data in table_data1:
                            data_list = data.split(':', 1)
                            if len(data_list) == 2:
                                key, value = data_list
                                formattted_key = key.lower().replace(" ", "").replace("/", "").replace(".", "")
                                table_data1_mapping[formattted_key] = value.strip()

                        page_info["Provider Identification"] = table_data1_mapping["custrefpono"]
                        page_info["Certificate No"] = table_data1_mapping["reportnumber"]
                        page_info["Previous Inspection"] = table_data1_mapping["dateofexamination"]



                        if quantity > 1:
                            identification_number_list = get_identification_number_list(id_numbers, quantity)

                        for identification_number in identification_number_list:
                            extraction_info[identification_number] = page_info

                        #
                    else:
                        print("No identification error")

        else:
            print("No verified company found")

    print(len(extraction_info.keys()))
    excel_management.create_excel(extraction_info, "../database/Centurion.xlsx", "Centurion", page_errors)


if __name__ == "__main__":
    extraction_centurion_pdf("../resources/CenturionLoft.pdf")









