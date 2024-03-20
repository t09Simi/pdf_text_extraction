import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side


column_mapping = {
    "Id Number": "A",
    "RFID": "B",
    "Item Category": "C",
    "Item Description": "D",
    "Model": "E",
    "SWL Value": "F",
    "SWL Unit": "G",
    "SWL Note": "H",
    "Manufacturer": "I",
    "Certificate No": "J",
    "Location": "K",
    "Detailed Location ": "L",
    "Previous Inspection": "M",
    "Next Inspection Due Date": "N",
    "Fit For Purpose Y/N": "O",
    "Status" : "P",
    "Provider Identification": "Q",
    "Errors": "R"
}


def create_excel(extracted_data: dict, filename: str, client: str, page_errors: dict):
    try:
        print("<--------------Creating new excel------------------>")
        workbook = openpyxl.Workbook()  # Create a new Workbook

        # Create a sheet for extracted data
        sheet_data = workbook.active
        sheet_data.title = "Extraction Data"  # Set sheet name

        sheet_data['A1'] = "Rig-Ware import v2"
        sheet_data['B1'] = client
        sheet_data['C1'] = "CreateLocations=No"

        # Write column headers for extracted data sheet
        for header, column in column_mapping.items():
            cell = sheet_data[column + '2']
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align the text
            cell.border = Border(bottom=Side(border_style='thin'))  # Add a thin border at the bottom

            # Adjust column width to fit the header text
            column_width = max(len(header), 10)  # Set a minimum width of 10 characters
            sheet_data.column_dimensions[column].width = column_width

        # Write extracted data to the worksheet
        for row_idx, (key, data) in enumerate(extracted_data.items(), start=3):
            sheet_data.cell(row=row_idx, column=1, value=key)  # Write key in the first column
            for cell_name, value in data.items():
                column_name = column_mapping.get(cell_name)
                if column_name:
                    sheet_data.cell(row=row_idx, column=ord(column_name) - 64, value=value)

        # Create a sheet for errors
        sheet_errors = workbook.create_sheet(title="Errors")  # Create a new worksheet
        sheet_errors.append(["Page No", "Error"])  # Write column headers

        # Write errors to the worksheet
        for key, value in page_errors.items():
            sheet_errors.append([key, value])  # Write key-value pairs as rows

        workbook.save(filename)  # Save the workbook with the provided filename
        workbook.close()
        print("<-------------- Excel created successfully ------------------>")
    except Exception as e:
        print(f"An error occurred in excel creation: {e}")


if __name__ == "__main__":
    create_excel(dict(), "output.xlsx", "Client Name", {"Page1": "Error1", "Page2": "Error2"})
